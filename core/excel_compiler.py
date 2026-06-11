import openpyxl
import io
import re
import sheets.main_sheet as main_sheet
import sheets.rent_calculation as rent_calculation
import sheets.lease_size as lease_size
import sheets.lease_rent as lease_rent
import sheets.capex_pm as capex_pm
import sheets.security_deposit as security_deposit

def compile_output_workbook(template_path, params):
    """Loads specimen template, populates the 5 sheets sequentially, and saves new workbook."""
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # Adjust all other sheet formulas in the template that reference the shifted rows in CAPEX and PM sheet
    # Run this FIRST so that any dynamically injected formulas during .inject() calls are not double-shifted
    N = len(params.get("Fitout Cost Breakdown", []))
    offset = 6 * (N - 3) if N > 3 else 0
    if offset != 0:
        pattern = re.compile(r"('CAPEX and PM'!|CAPEX and PM!)(\$?[A-Z]+)(\$?)([0-9]+)")
        def repl(match):
            prefix = match.group(1)
            col = match.group(2)
            abs_sign = match.group(3)
            row_str = match.group(4)
            row_num = int(row_str)
            if row_num >= 21:
                return f"{prefix}{col}{abs_sign}{row_num + offset}"
            return match.group(0)
            
        for sheet_name in wb.sheetnames:
            if sheet_name != "CAPEX and PM":
                ws = wb[sheet_name]
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if isinstance(val, str) and val.startswith("="):
                            new_val = pattern.sub(repl, val)
                            if new_val != val:
                                ws.cell(row=r, column=c, value=new_val)
                                
        # Shift local formulas inside CAPEX and PM sheet itself
        ws_cp = wb["CAPEX and PM"]
        local_pattern = re.compile(r"(?<![A-Za-z0-9_!])(\$?[A-Z]+)(\$?)([0-9]+)")
        def local_repl(match):
            col_part = match.group(1)
            abs_sign = match.group(2)
            row_str = match.group(3)
            row_num = int(row_str)
            if row_num >= 21:
                return f"{col_part}{abs_sign}{row_num + offset}"
            return match.group(0)
            
        for r in range(1, ws_cp.max_row + 1):
            for c in range(1, ws_cp.max_column + 1):
                val = ws_cp.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    new_val = local_pattern.sub(local_repl, val)
                    if new_val != val:
                        ws_cp.cell(row=r, column=c, value=new_val)
    
    # Populate Main sheet
    main_sheet.inject(wb["Main"], params)
    
    # Populate Rent Calculation sheet
    rent_calculation.inject(wb["Rent Calculation"], params)
    
    # Populate Lease Size sheet
    lease_size.inject(wb["Lease Size"], params)
    
    # Populate Lease Rent sheet
    lease_rent.inject(wb["Lease Rent"], params)
    
    # Populate CAPEX and PM sheet
    capex_pm.inject(wb["CAPEX and PM"], params)
    
    # Populate Security Deposit sheet
    security_deposit.inject(wb["Security Deposit"], params)
                                
    for ws in wb.worksheets:
        if ws.sheet_properties:
            ws.sheet_properties.tabColor = None
            
    # Save to a memory stream for download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
