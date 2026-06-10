import streamlit as st
import pandas as pd
import openpyxl
import datetime
import os
import io

import importlib

import sheets.main_sheet as main_sheet
import sheets.rent_calculation as rent_calculation
import sheets.lease_size as lease_size
import sheets.lease_rent as lease_rent
import sheets.capex_pm as capex_pm
import sheets.security_deposit as security_deposit
import sheets.input_parser as input_parser

# Force reload sub-modules in Streamlit to ensure edits are picked up immediately
importlib.reload(main_sheet)
importlib.reload(rent_calculation)
importlib.reload(lease_size)
importlib.reload(lease_rent)
importlib.reload(capex_pm)
importlib.reload(security_deposit)
importlib.reload(input_parser)

from sheets.capex_pm import calculate_fy_months
from sheets.input_parser import parse_input_xlsx


st.set_page_config(
    page_title="Siemens - Rental Automation System",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- TEMPLATE UTILITIES ---
def load_template(filename, directory="templates"):
    """Reads raw file contents from templates directory for dynamic HTML/CSS injection."""
    path = os.path.join(directory, filename)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except Exception:
            return ""
    return ""

# Inject css file from templates
style_css = load_template("style.css")
if style_css:
    st.markdown(f"<style>{style_css}</style>", unsafe_allow_html=True)

# Inject background image
import base64
bg_path = os.path.join(os.path.dirname(__file__), "bg.png")
if os.path.exists(bg_path):
    try:
        with open(bg_path, "rb") as f:
            bg_data = f.read()
        bg_base64 = base64.b64encode(bg_data).decode()
        bg_css = f"""
        <style>
        .stApp {{
            background-image: url("data:image/png;base64,{bg_base64}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}
        </style>
        """
        st.markdown(bg_css, unsafe_allow_html=True)
    except Exception:
        pass




def compile_output_workbook(template_path, params):
    """Loads specimen template, populates the 5 sheets sequentially, and saves new workbook."""
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # Adjust all other sheet formulas in the template that reference the shifted rows in CAPEX and PM sheet
    # Run this FIRST so that any dynamically injected formulas during .inject() calls are not double-shifted
    N = len(params.get("Fitout Cost Breakdown", []))
    offset = 6 * (N - 3) if N > 3 else 0
    if offset != 0:
        import re
        pattern = re.compile(r"('CAPEX and PM'!|CAPEX and PM!)(\$?[A-Z]+)(\$?)([0-9]+)")
        def repl(match):
            prefix = match.group(1)
            col = match.group(2)
            abs_sign = match.group(3)
            row_str = match.group(4)
            row_num = int(row_str)
            if row_num >= 21:
                return f"{prefix}{col}{abs_sign}{row_num + offset}"
            return match.group(0)
            
        for sheet_name in wb.sheetnames:
            if sheet_name != "CAPEX and PM":
                ws = wb[sheet_name]
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if isinstance(val, str) and val.startswith("="):
                            new_val = pattern.sub(repl, val)
                            if new_val != val:
                                ws.cell(row=r, column=c, value=new_val)
                                
        # Shift local formulas inside CAPEX and PM sheet itself
        ws_cp = wb["CAPEX and PM"]
        local_pattern = re.compile(r"(?<![A-Za-z0-9_!])(\$?[A-Z]+)(\$?)([0-9]+)")
        def local_repl(match):
            col_part = match.group(1)
            abs_sign = match.group(2)
            row_str = match.group(3)
            row_num = int(row_str)
            if row_num >= 21:
                return f"{col_part}{abs_sign}{row_num + offset}"
            return match.group(0)
            
        for r in range(1, ws_cp.max_row + 1):
            for c in range(1, ws_cp.max_column + 1):
                val = ws_cp.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    new_val = local_pattern.sub(local_repl, val)
                    if new_val != val:
                        ws_cp.cell(row=r, column=c, value=new_val)
    
    # Populate Main sheet
    main_sheet.inject(wb["Main"], params)
    
    # Populate Rent Calculation sheet
    rent_calculation.inject(wb["Rent Calculation"], params)
    
    # Populate Lease Size sheet
    lease_size.inject(wb["Lease Size"], params)
    
    # Populate Lease Rent sheet
    lease_rent.inject(wb["Lease Rent"], params)
    
    # Populate CAPEX and PM sheet
    capex_pm.inject(wb["CAPEX and PM"], params)
    
    # Populate Security Deposit sheet
    security_deposit.inject(wb["Security Deposit"], params)
                                
    for ws in wb.worksheets:
        if ws.sheet_properties:
            ws.sheet_properties.tabColor = None
            
    # Save to a memory stream for download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- STREAMLIT UI LAYOUT & CONTROLS ---

# 1. Render custom premium header
header_html = load_template("header.html")
if header_html:
    st.markdown(header_html, unsafe_allow_html=True)
else:
    st.title("🏢 Rental Automation System")

# 2. Sidebar Ingestion Panel
st.sidebar.markdown('<p style="font-size: 1.3rem; font-weight:700; color:#f8fafc; margin-bottom: 15px;">📁 Data Ingestion</p>', unsafe_allow_html=True)
uploaded_file = st.sidebar.file_uploader("Upload Lease Parameter Workbook (.xlsx)", type=["xlsx"], help="Upload an Excel sheet (.xlsx) containing main lease configuration parameters.")

input_template_path = os.path.join(os.path.dirname(__file__), "artifacts", "input_template.xlsx")
if os.path.exists(input_template_path):
    try:
        with open(input_template_path, "rb") as f:
            template_bytes = f.read()
        st.sidebar.download_button(
            label="📥 Download Input Template",
            data=template_bytes,
            file_name="input_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Download the blank template sheet to fill in lease parameters."
        )
    except Exception as e:
        st.sidebar.error(f"Error loading template: {e}")

# Default fallback values matching Specimen
default_params = {
    "REU Name": "IN-CHEK2",
    "Building Name": "SKCL Tech Park",
    "City": "Chennai",
    "Country": "India",
    "Currency": "INR",
    "Lease Type": "Office",
    "Chargeable Area Sqft": 9515.0,
    "Agreement Start Date": datetime.date(2026, 4, 1),
    "Agreement End Date": datetime.date(2031, 3, 31),
    "Rent Per Sqft": 120.0,
    "Quoted CAM": 15.48,
    "Escalation %": 0.15,
    "Escalation Frequency Months": 36,
    "CAM Escalation %": 0.05,
    "CAM Escalation Frequency Months": 12,
    "4 Wheeler Slots": 80,
    "4 Wheeler Rate": 1500.0,
    "2 Wheeler Slots": 50,
    "2 Wheeler Rate": 1000.0,
    "Parking Escalation %": 0.15,
    "Parking Escalation Frequency Months": 36,
    "Security Deposit Amount": 11418000,
    "Addnl.Deposit -energy(Refundable)": 500000,
    "Fitout Cost": 34000000,
    "Fitout Cost Breakdown": [14000000.0, 5000000.0, 15000000.0],
    "Fitout Useful Lives": [72, 30, 48],
    "Fitout Active Months Breakdown": [
        {2026: 9, 2027: 12, 2028: 12, 2029: 12, 2030: 12, 2031: 3},
        {2026: 9, 2027: 12, 2028: 9},
        {2026: 9, 2027: 12, 2028: 12, 2029: 12, 2030: 3}
    ],
    "Capex Schedule": {2026: 14000000.0, 2027: 12000000.0, 2028: 5000000.0, 2029: 6000000.0, 2030: 6000000.0},
    "Capex Useful Lives": {2026: 72, 2027: 60, 2028: 48, 2029: 36, 2030: 24},
    "PM Cost Over Lease": 2500000.0,
    "PM Schedule": {2026: 1500000.0, 2027: 200000.0, 2028: 200000.0, 2029: 200000.0, 2030: 200000.0, 2031: 200000.0},
    "Cost of Capital": 0.105,
    "Discount Rate": 0.08,
    "Incremental Borrowing Rate": 0.08,
    "Imputed Interest Rate": 0.0711,
    "Ready Reckoner Rate": 15000.0,
    "Exchange Rate": 105.02,
    "Incremental Restoration Cost Sqft": 82.6,
    "Opex Others Per Month": 654.0,
    "Opex II Per Month": 0.0,
    "Lease Term Months": 60
}

params = default_params
parse_error_msg = None
upload_occurred = False

# Ensure upload_counter is in session_state
if "upload_counter" not in st.session_state:
    st.session_state["upload_counter"] = 0

if uploaded_file is not None:
    upload_occurred = True
    file_name = uploaded_file.name
    # If the file uploaded is different from the last one, increment counter to reset keys
    if st.session_state.get("last_uploaded_file") != file_name:
        st.session_state["last_uploaded_file"] = file_name
        st.session_state["upload_counter"] += 1
        
    parsed_vals, err = parse_input_xlsx(uploaded_file)
    if err:
        parse_error_msg = err
    else:
        params = parsed_vals
else:
    # If file was cleared, increment counter to reset back to default params
    if "last_uploaded_file" in st.session_state:
        del st.session_state["last_uploaded_file"]
        st.session_state["upload_counter"] += 1

key_suffix = f"_up_{st.session_state['upload_counter']}"

# 3. Sidebar Override & Configuration Panel
st.sidebar.markdown('<p style="font-size: 1.25rem; font-weight:600; color:#cbd5e1; margin-top: 20px;">⚙️ Parameter Configuration</p>', unsafe_allow_html=True)

# Expander 1: Real Estate & Lease Properties
with st.sidebar.expander("▶ Real Estate & Lease Properties", expanded=True):
    reu_name = st.text_input("REU Name", value=params["REU Name"], key=f"reu_name{key_suffix}", help="Unique identifier code for the Real Estate Unit.")
    bldg_name = st.text_input("Building Name", value=params["Building Name"], key=f"bldg_name{key_suffix}", help="The name of the building/facility where the premises are leased.")
    city = st.text_input("City", value=params["City"], key=f"city{key_suffix}", help="The city location of the leased property.")
    area = st.number_input("Chargeable Area (sq ft)", value=int(params["Chargeable Area Sqft"]), key=f"area{key_suffix}", step=100, help="The chargeable/rentable area in square feet.")
    
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        start_date = st.date_input("Start Date", value=params["Agreement Start Date"], key=f"start_date{key_suffix}", help="The official lease agreement commencement date.")
    with col_d2:
        end_date = st.date_input("End Date", value=params["Agreement End Date"], key=f"end_date{key_suffix}", help="The official lease agreement expiration date.")
        
    rent_sqft = st.number_input("Quoted Rent (per sq ft/month)", value=float(params["Rent Per Sqft"]), key=f"rent_sqft{key_suffix}", step=1.0, help="The starting rental rate per square foot per month.")
    cam_sqft = st.number_input("Quoted CAM (per sq ft/month)", value=float(params["Quoted CAM"]), key=f"cam_sqft{key_suffix}", step=0.1, help="The starting Common Area Maintenance (CAM) rate per square foot per month.")
    
    rent_esc = st.slider("Rent Escalation %", min_value=0.0, max_value=0.5, value=float(params["Escalation %"]), key=f"rent_esc{key_suffix}", step=0.01, help="The rental rate escalation percentage (e.g. 0.15 = 15%).")
    rent_esc_freq = st.number_input("Rent Escalation Freq (Months)", value=int(params["Escalation Frequency Months"]), key=f"rent_esc_freq{key_suffix}", step=12, help="Interval in months between rent escalations (typically 36 months).")
    cam_esc = st.slider("CAM Escalation %", min_value=0.0, max_value=0.2, value=float(params["CAM Escalation %"]), key=f"cam_esc{key_suffix}", step=0.01, help="Escalation percentage applied to CAM rates.")
    cam_esc_freq = st.number_input("CAM Escalation Freq (Months)", value=int(params.get("CAM Escalation Frequency Months", 12)), key=f"cam_esc_freq{key_suffix}", step=12, help="Interval in months between CAM rate escalations (typically 12 months).")
    
    sec_deposit = st.number_input("Security Deposit Amount (INR)", value=int(params["Security Deposit Amount"]), key=f"sec_deposit{key_suffix}", step=50000, help="Interest-free refundable security deposit amount.")
    energy_dep = st.number_input("Energy Deposit Amount (INR)", value=int(params["Addnl.Deposit -energy(Refundable)"]), key=f"energy_dep{key_suffix}", step=10000, help="Additional refundable security deposit specifically for power/utilities.")

# Expander 1.5: Parking Configuration
with st.sidebar.expander("▶ Parking Configuration"):
    col_p1, col_p2 = st.columns(2)
    with col_p1:
        four_w_slots = st.number_input("4 Wheeler Slots", value=int(params.get("4 Wheeler Slots", 80)), key=f"four_w_slots{key_suffix}", step=1, help="Number of chargeable 4-wheeler parking slots.")
    with col_p2:
        four_w_rate = st.number_input("4 Wheeler Rate (per mo)", value=float(params.get("4 Wheeler Rate", 1500.0)), key=f"four_w_rate{key_suffix}", step=100.0, help="Monthly rate per 4-wheeler slot.")
    col_p3, col_p4 = st.columns(2)
    with col_p3:
        two_w_slots = st.number_input("2 Wheeler Slots", value=int(params.get("2 Wheeler Slots", 50)), key=f"two_w_slots{key_suffix}", step=1, help="Number of chargeable 2-wheeler parking slots.")
    with col_p4:
        two_w_rate = st.number_input("2 Wheeler Rate (per mo)", value=float(params.get("2 Wheeler Rate", 1000.0)), key=f"two_w_rate{key_suffix}", step=100.0, help="Monthly rate per 2-wheeler slot.")
        
    col_p5, col_p6 = st.columns(2)
    with col_p5:
        parking_esc = st.slider("Parking Escalation %", min_value=0.0, max_value=0.5, value=float(params.get("Parking Escalation %", params.get("Escalation %", 0.15))), key=f"parking_esc{key_suffix}", step=0.01, help="The parking rate escalation percentage.")
    with col_p6:
        parking_esc_freq = st.number_input("Parking Escalation Freq (Months)", value=int(params.get("Parking Escalation Frequency Months", params.get("Escalation Frequency Months", 36))), key=f"parking_esc_freq{key_suffix}", step=1, help="Interval in months between parking rate escalations.")

# Expander 1.6: OpEx Configuration
with st.sidebar.expander("▶ OpEx Configuration"):
    opex_others = st.number_input("Opex I (Others per month)", value=float(params.get("Opex Others Per Month", 654.0)), key=f"opex_others{key_suffix}", step=10.0, help="Opex Others monthly amount.")
    opex_ii = st.number_input("Opex II (per month)", value=float(params.get("Opex II Per Month", 0.0)), key=f"opex_ii{key_suffix}", step=10.0, help="Opex II monthly amount (OpEx Add-on).")

# Expander 2: CAPEX & PM Investment Schedule
with st.sidebar.expander("▶ CAPEX & PM Schedule"):
    st.markdown("**Fitout Investments (Breakdown)**")
    fitout_breakdown = params.get("Fitout Cost Breakdown", [14000000.0, 5000000.0, 15000000.0])
    fitout_lifes = params.get("Fitout Useful Lives", [72, 30, 48])
    
    # Filter out zeros from breakdown to determine configured phases
    non_zero_fitouts = [f for f in fitout_breakdown if f > 0]
    default_num_fitouts = max(1, len(non_zero_fitouts))
    max_fitouts = max(8, default_num_fitouts)
    num_fitouts = st.selectbox("Number of Fitout Phases", list(range(1, max_fitouts + 1)), index=default_num_fitouts - 1, key=f"num_fitouts{key_suffix}", help="Select the number of dynamic fitout phases.")
    
    fitout_costs = []
    fitout_useful_lives = []
    fitout_active_months_breakdown = []
    
    # Calculate years of the lease
    lease_years = list(range(start_date.year, end_date.year + 1))
    from sheets.capex_pm import calculate_fy_months
    
    term_months = round((end_date - start_date).days / 30.4167)
    parsed_active_months_list = params.get("Fitout Active Months Breakdown", [])
    
    for i in range(num_fitouts):
        col1, col2 = st.columns(2)
        default_c = fitout_breakdown[i] if i < len(fitout_breakdown) else 0.0
        default_l = fitout_lifes[i] if i < len(fitout_lifes) else 12
        with col1:
            c = st.number_input(f"Fitout Phase {i+1} Cost", value=int(default_c), key=f"fitout_cost_{i}{key_suffix}", step=100000, help=f"Capital expenditure for phase {i+1} of fitout works.")
        with col2:
            l = st.number_input(f"Phase {i+1} Useful Life (mo)", value=int(default_l), key=f"fitout_life_{i}{key_suffix}", step=6, help=f"Amortization period for phase {i+1} in months.")
            
        fitout_costs.append(float(c))
        fitout_useful_lives.append(int(l))
        
        # Expandable year-by-year active months customization
        with st.expander(f"Phase {i+1} Year-by-Year Months Active"):
            default_m_dist = calculate_fy_months(start_date, term_months, int(l))
            phase_m_dict = parsed_active_months_list[i] if i < len(parsed_active_months_list) else {}
            
            phase_m_custom = {}
            for yr in lease_years:
                default_m = phase_m_dict.get(yr, default_m_dist.get(yr, 0))
                m = st.number_input(f"FY{yr} Active Months", min_value=0, max_value=12, value=int(default_m), key=f"fitout_{i}_m_{yr}{key_suffix}")
                phase_m_custom[yr] = int(m)
            fitout_active_months_breakdown.append(phase_m_custom)
        
    fitout_total = sum(fitout_costs)
    st.info(f"Total Fitout Cost: {params['Currency']} {fitout_total:,.2f}")
    
    st.markdown("**Capex Investments by Year**")
    capex_dict = params.get("Capex Schedule", {})
    # Determine non-zero elements
    non_zero_capex_years = [y for y, val in capex_dict.items() if val > 0]
    start_year = start_date.year
    
    # Determine default number of years based on max year configured
    if non_zero_capex_years:
        default_num_capex = max(1, max(non_zero_capex_years) - start_year + 1)
    else:
        default_num_capex = 5
        
    max_capex = max(8, default_num_capex)
    num_capex = st.selectbox("Number of Capex Years", list(range(1, max_capex + 1)), index=default_num_capex - 1, key=f"num_capex{key_suffix}", help="Select the number of years for CAPEX injection.")
    
    capex_schedule = {}
    capex_useful_lives = {}
    capex_lives_dict = params.get("Capex Useful Lives", {})
    term_months = round((end_date - start_date).days / 30.4167)
    
    for i in range(num_capex):
        yr = start_year + i
        default_val = capex_dict.get(yr, 0.0)
        
        # Calculate default life (remaining lease term)
        if i == 0:
            default_life = term_months
        else:
            first_year_months = 12 - start_date.month + 1
            elapsed = first_year_months + 12 * (i - 1)
            default_life = max(0, term_months - elapsed)
            
        default_l = capex_lives_dict.get(yr, default_life)
        
        col1, col2 = st.columns(2)
        with col1:
            val = st.number_input(f"Capex FY{yr} Cost", value=int(default_val), key=f"capex_cost_{i}{key_suffix}", step=100000, help=f"Estimated Capital Expenditures for Fiscal Year {yr}.")
        with col2:
            life = st.number_input(f"FY{yr} Useful Life (mo)", value=int(default_l), key=f"capex_life_{i}{key_suffix}", step=12, help=f"Amortization period for CAPEX in Fiscal Year {yr} in months.")
            
        capex_useful_lives[yr] = int(life)
        if val > 0:
            capex_schedule[yr] = float(val)
            
    capex_total = sum(capex_schedule.values())
    st.info(f"Total Capex Cost: {params['Currency']} {capex_total:,.2f}")
    
    st.markdown("**Maintenance (PM) Schedule**")
    pm_dict = params.get("PM Schedule", {})
    non_zero_pm_years = [y for y, val in pm_dict.items() if val > 0]
    
    if non_zero_pm_years:
        default_num_pm = max(1, max(non_zero_pm_years) - start_year + 1)
    else:
        default_num_pm = 6
        
    max_pm = max(10, default_num_pm)
    num_pm = st.selectbox("Number of PM Years", list(range(1, max_pm + 1)), index=default_num_pm - 1, key=f"num_pm{key_suffix}", help="Select the number of years for Preventive Maintenance schedule.")
    
    pm_schedule = {}
    for i in range(num_pm):
        yr = start_year + i
        default_val = pm_dict.get(yr, 0.0)
        val = st.number_input(f"PM FY{yr}", value=int(default_val), key=f"pm_cost_{i}{key_suffix}", step=50000, help=f"Preventive Maintenance budget allocation for Fiscal Year {yr}.")
        if val > 0:
            pm_schedule[yr] = float(val)
            
    pm_cost_total = sum(pm_schedule.values())
    st.info(f"Total PM Cost: {params['Currency']} {pm_cost_total:,.2f}")

# Expander 3: Financial Rates & ARO Restoration
with st.sidebar.expander("▶ Rates & ARO Restoration"):
    wacc = st.number_input("Cost of Capital (WACC) %", value=float(params["Cost of Capital"]), format="%.4f", step=0.001, key=f"wacc{key_suffix}", help="Weighted Average Cost of Capital used for investment analysis and discounting cash flows.")
    imputed_rate = st.number_input("Imputed Interest Rate %", value=float(params["Imputed Interest Rate"]), format="%.4f", step=0.001, key=f"imputed_rate{key_suffix}", help="The implicit rate of interest in the lease, or estimated discount rate.")
    aro_rate = st.number_input("Restoration Cost per Sq ft (ARO)", value=float(params["Incremental Restoration Cost Sqft"]), key=f"aro_rate{key_suffix}", step=5.0, help="Estimated asset restoration cost per square foot at lease end (Asset Retirement Obligation).")
    exchange_rate = st.number_input("Forex Rate (INR/Euro)", value=float(params["Exchange Rate"]), key=f"exchange_rate{key_suffix}", step=0.1, help="The foreign currency exchange rate (INR per 1 Euro) for reporting.")

# Pack overrides into active UI parameters
ui_params = {
    "REU Name": reu_name,
    "Building Name": bldg_name,
    "City": city,
    "Country": params["Country"],
    "Currency": params["Currency"],
    "Lease Type": params["Lease Type"],
    "Chargeable Area Sqft": float(area),
    "Agreement Start Date": start_date,
    "Agreement End Date": end_date,
    "Rent Per Sqft": float(rent_sqft),
    "Quoted CAM": float(cam_sqft),
    "Escalation %": float(rent_esc),
    "Escalation Frequency Months": int(rent_esc_freq),
    "CAM Escalation %": float(cam_esc),
    "CAM Escalation Frequency Months": int(cam_esc_freq),
    "Security Deposit Amount": float(sec_deposit),
    "Addnl.Deposit -energy(Refundable)": float(energy_dep),
    "4 Wheeler Slots": int(four_w_slots),
    "4 Wheeler Rate": float(four_w_rate),
    "2 Wheeler Slots": int(two_w_slots),
    "2 Wheeler Rate": float(two_w_rate),
    "Parking Escalation %": float(parking_esc),
    "Parking Escalation Frequency Months": int(parking_esc_freq),
    
    "Fitout Cost": float(fitout_total),
    "Fitout Cost Breakdown": fitout_costs,
    "Fitout Useful Lives": fitout_useful_lives,
    "Fitout Active Months Breakdown": fitout_active_months_breakdown,
    "Capex Schedule": capex_schedule,
    "Capex Useful Lives": capex_useful_lives,
    "PM Cost Over Lease": float(pm_cost_total),
    "PM Schedule": pm_schedule,
    
    "Cost of Capital": float(wacc),
    "Discount Rate": float(params.get("Discount Rate", params.get("Incremental Borrowing Rate", 0.08))),
    "Incremental Borrowing Rate": float(params.get("Incremental Borrowing Rate", params.get("Discount Rate", 0.08))),
    "Imputed Interest Rate": float(imputed_rate),
    "Ready Reckoner Rate": float(params.get("Ready Reckoner Rate", 15000.0)),
    "Exchange Rate": float(exchange_rate),
    "Incremental Restoration Cost Sqft": float(aro_rate),
    "Opex Others Per Month": float(opex_others),
    "Opex II Per Month": float(opex_ii),
    "Lease Term Months": round((end_date - start_date).days / 30.4167)
}


# --- LIVE SIMULATION MODEL W wake-up ---
capex_pm_df = capex_pm.simulate(ui_params)
sd_results = security_deposit.simulate(ui_params)
rent_calc_results = rent_calculation.simulate(ui_params)
lease_size_df, npv_value = lease_size.simulate(ui_params, capex_pm_df)



# --- DYNAMIC STATUS CARD SECTION ---
if parse_error_msg:
    error_tpl = load_template("error_card.html")
    if error_tpl:
        st.markdown(error_tpl.format(error_message=f"Failed to parse sheet: {parse_error_msg}"), unsafe_allow_html=True)
    else:
        st.error(f"Failed to parse sheet: {parse_error_msg}")
elif upload_occurred:
    success_tpl = load_template("success_card.html")
    if success_tpl:
        total_sd = ui_params["Security Deposit Amount"] + ui_params["Addnl.Deposit -energy(Refundable)"]
        duration_yrs = ui_params["Lease Term Months"] / 12.0
        st.markdown(success_tpl.format(
            reu_name=ui_params["REU Name"],
            area=f"{int(ui_params['Chargeable Area Sqft']):,}",
            duration=f"{duration_yrs:.2f}",
            currency=ui_params["Currency"],
            total_deposit=f"{total_sd:,.2f}"
        ), unsafe_allow_html=True)
    else:
        st.success(f"Workbook parameters for {ui_params['REU Name']} compiled successfully!")
else:
    info_tpl = load_template("info_card.html")
    if info_tpl:
        st.markdown(info_tpl, unsafe_allow_html=True)


# --- LIVE KPI DASHBOARD (HTML GRID) ---
total_rent_cam = ui_params["Rent Per Sqft"] + ui_params["Quoted CAM"]
duration_yrs = ui_params["Lease Term Months"] / 12.0
total_sd_amount = ui_params["Security Deposit Amount"] + ui_params["Addnl.Deposit -energy(Refundable)"]
total_fitout = ui_params["Fitout Cost"]
total_capex = sum(ui_params["Capex Schedule"].values())
total_pm = ui_params["PM Cost Over Lease"]

net_rent_1 = rent_calc_results["Net Rent I (Standard)"]
net_rent_2 = rent_calc_results["Net Rent II (Refinancing)"]
opex_1 = rent_calc_results["Opex Others Per Month"]
opex_2 = rent_calc_results["Opex II Per Month"]
total_occupancy_cost = rent_calc_results["Total Occupancy Cost"]

metrics_html = f"""
<div class="stats-grid">
    <div class="stat-cell">
        <span class="stat-label">Chargeable Area</span>
        <span class="stat-value">{int(ui_params['Chargeable Area Sqft']):,} sqft</span>
        <span class="stat-meta">≈ {ui_params['Chargeable Area Sqft']/10.764:,.2f} sq.m.</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Initial Rent + CAM Rate</span>
        <span class="stat-value">{ui_params['Currency']} {total_rent_cam:.2f}</span>
        <span class="stat-meta">Rent: {ui_params['Rent Per Sqft']:.1f} | CAM: {ui_params['Quoted CAM']:.2f}</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Lease Duration</span>
        <span class="stat-value">{duration_yrs:.2f} yrs</span>
        <span class="stat-meta">{ui_params['Agreement Start Date'].strftime('%b %d, %Y')} to {ui_params['Agreement End Date'].strftime('%b %d, %Y')}</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Total Fitout Cost</span>
        <span class="stat-value">{ui_params['Currency']} {total_fitout/1000000:,.2f} M</span>
        <span class="stat-meta">{len(ui_params['Fitout Cost Breakdown'])} Phases Configured</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Total CAPEX</span>
        <span class="stat-value">{ui_params['Currency']} {total_capex/1000000:,.2f} M</span>
        <span class="stat-meta">{len(ui_params['Capex Schedule'])} Years Scheduled</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Total PM Cost</span>
        <span class="stat-value">{ui_params['Currency']} {total_pm/1000000:,.2f} M</span>
        <span class="stat-meta">{len(ui_params['PM Schedule'])} Years Scheduled</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Project NPV</span>
        <span class="stat-value">€ {npv_value:,.2f} M</span>
        <span class="stat-meta">WACC: {ui_params['Cost of Capital']*100:.2f}% | Forex: {ui_params['Exchange Rate']:.2f}</span>
    </div>
    <div class="stat-cell">
        <span class="stat-label">Total Occupancy Cost</span>
        <span class="stat-value">{ui_params['Currency']} {total_occupancy_cost:,.2f}</span>
        <div class="stat-sub-grid">
            <div class="stat-sub-item">
                <span class="stat-sub-label">Net Rent 1</span>
                <span class="stat-sub-value">{net_rent_1:,.2f}</span>
            </div>
            <div class="stat-sub-item">
                <span class="stat-sub-label">Net Rent 2</span>
                <span class="stat-sub-value">{net_rent_2:,.2f}</span>
            </div>
            <div class="stat-sub-item">
                <span class="stat-sub-label">Opex 1</span>
                <span class="stat-sub-value">{opex_1:,.2f}</span>
            </div>
            <div class="stat-sub-item">
                <span class="stat-sub-label">Opex 2</span>
                <span class="stat-sub-value">{opex_2:,.2f}</span>
            </div>
        </div>
    </div>
</div>
"""

st.markdown(metrics_html, unsafe_allow_html=True)


# --- EXCEL WORKBOOK COMPILER GENERATION ---
template_file_path = os.path.join(os.path.dirname(__file__), "artifacts", "Rental Specimen.xlsx")

if os.path.exists(template_file_path):
    try:
        output_stream = compile_output_workbook(template_file_path, ui_params)
        
        st.download_button(
            label="💾 Generate and Download Excel Workbook",
            data=output_stream,
            file_name=f"rental_workbook_{ui_params['REU Name']}_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    except Exception as e:
        st.error(f"Excel template compilation failed: {e}")
else:
    st.warning(f"Excel template specimen not found in path: {template_file_path}")

