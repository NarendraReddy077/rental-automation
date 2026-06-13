import io
import re
import pandas as pd
import openpyxl

def parse_extra_sheets(raw_data, params):
    """Uses openpyxl to parse extra sheets (Lease Rent, Rent Calculation, CAPEX and PM).
    
    Modifies params in-place for parking rates/slots and exchange rate.
    Returns:
        fitout_breakdown (list)
        fitout_lifes (list)
        fitout_active_months_breakdown (list of dicts)
        capex_sched (dict)
        capex_lives (dict)
        pm_sched (dict)
        has_capex_pm_sheet (bool)
    """
    fitout_breakdown = []
    fitout_lifes = []
    fitout_active_months_breakdown = []
    capex_sched = {}
    capex_lives = {}
    pm_sched = {}
    
    wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
    
    if "Lease Rent" in wb.sheetnames:
        ws_lr = wb["Lease Rent"]
        val_c22 = ws_lr["C22"].value
        val_d22 = ws_lr["D22"].value
        val_c23 = ws_lr["C23"].value
        val_d23 = ws_lr["D23"].value
        if val_c22 is not None:
            try:
                params["4 Wheeler Rate"] = float(val_c22)
            except Exception:
                pass
        if val_d22 is not None:
            try:
                params["4 Wheeler Slots"] = int(float(val_d22))
            except Exception:
                pass
        if val_c23 is not None:
            try:
                params["2 Wheeler Rate"] = float(val_c23)
            except Exception:
                pass
        if val_d23 is not None:
            try:
                params["2 Wheeler Slots"] = int(float(val_d23))
            except Exception:
                pass

    if "Rent Calculation" in wb.sheetnames:
        ws_rc = wb["Rent Calculation"]
        val_j5 = ws_rc["J5"].value
        if val_j5 is not None:
            try:
                params["Exchange Rate"] = float(val_j5)
            except Exception:
                pass

    has_capex_pm_sheet = "CAPEX and PM" in wb.sheetnames
    if has_capex_pm_sheet:
        ws = wb["CAPEX and PM"]
        
        # 1. Parse fitout costs, useful lives and active months dynamically (up to 8 phases)
        for k in range(8):
            r_c = 6 + 6 * k
            r_m = 3 + 6 * k
            lbl = ws.cell(row=r_c, column=1).value
            if lbl == "Investment Cost":
                cost_val = ws.cell(row=r_c, column=2).value
                life_val_months = ws.cell(row=r_m, column=5).value
                
                cost = float(cost_val) if isinstance(cost_val, (int, float)) else 0.0
                life = int(life_val_months) if isinstance(life_val_months, (int, float)) else 72
                
                fitout_breakdown.append(cost)
                fitout_lifes.append(life)
                
                # Parse active months for this phase year-by-year (columns F to O)
                phase_m = {}
                for c in range(6, 16):
                    yr_val = ws.cell(row=2, column=c).value
                    m_val = ws.cell(row=r_m, column=c).value
                    try:
                        yr = int(float(str(yr_val).strip()))
                    except (ValueError, TypeError):
                        continue
                    phase_m[yr] = int(m_val) if isinstance(m_val, (int, float)) else 0
                fitout_active_months_breakdown.append(phase_m)
        
        # 2. Parse Capex schedule and useful lives
        r_capex_header = 36
        for r in range(1, 100):
            if ws.cell(row=r, column=1).value == "Investment name" and ws.cell(row=r, column=2).value == "Capex":
                r_capex_header = r
                break
        
        r_capex_val = r_capex_header + 2
        for c in range(6, 16):
            yr_val = ws.cell(row=r_capex_header, column=c).value
            val = ws.cell(row=r_capex_val, column=c).value
            tranche_row = r_capex_header + 4 + 5 * (c - 6)
            life_val = ws.cell(row=tranche_row, column=5).value
            try:
                yr = int(float(str(yr_val).strip()))
            except (ValueError, TypeError):
                continue
            try:
                val_num = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                val_num = 0.0
            if val_num > 0:
                capex_sched[yr] = val_num
                try:
                    capex_lives[yr] = int(life_val) if life_val is not None else 0
                except (ValueError, TypeError):
                    pass
                    
        # 3. Parse PM schedule from PM row
        r_pm = 95
        for r in range(1, 150):
            if ws.cell(row=r, column=1).value == "Basic and project maintenance":
                r_pm = r
                break
        for c in range(6, 16):
            yr_val = ws.cell(row=2, column=c).value
            val = ws.cell(row=r_pm, column=c).value
            try:
                yr = int(float(str(yr_val).strip()))
            except (ValueError, TypeError):
                continue
            try:
                val_num = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                val_num = 0.0
            if val_num > 0:
                pm_sched[yr] = val_num

    return (
        fitout_breakdown,
        fitout_lifes,
        fitout_active_months_breakdown,
        capex_sched,
        capex_lives,
        pm_sched,
        has_capex_pm_sheet
    )

def parse_schedules_from_main_fallback(raw_params):
    """Regex parses fitout cost breakdown, Capex schedule, and PM schedule from raw_params."""
    fitout_breakdown = []
    capex_sched = {}
    pm_sched = {}
    
    # 1. Fitout breakdown costs
    fitout_costs_dict = {}
    for key, val in raw_params.items():
        key_clean = key.lower()
        if "fitout" in key_clean and "phase" in key_clean:
            m = re.search(r"\d+", key)
            if m:
                try:
                    phase_num = int(m.group(0))
                    if val is not None and not pd.isna(val):
                        fitout_costs_dict[phase_num] = float(val)
                except Exception:
                    pass
    for p_num in sorted(fitout_costs_dict.keys()):
        fitout_breakdown.append(fitout_costs_dict[p_num])
                
    # 2. Capex schedule
    for key, val in raw_params.items():
        key_clean = key.lower().replace(" ", "")
        if "capex" in key_clean:
            m = re.search(r"\b(20\d{2})\b|(?<=fy)(20\d{2})", key_clean)
            if not m:
                m = re.search(r"\d{4}", key)
            if m:
                try:
                    yr = int(m.group(0))
                    if val is not None and not pd.isna(val):
                        capex_sched[yr] = float(val)
                except Exception:
                    pass
                
    # 3. PM schedule
    for key, val in raw_params.items():
        key_clean = key.lower().replace(" ", "")
        if "pm" in key_clean or "maintenance" in key_clean:
            m = re.search(r"\b(20\d{2})\b|(?<=fy)(20\d{2})", key_clean)
            if not m:
                m = re.search(r"\d{4}", key)
            if m:
                try:
                    yr = int(m.group(0))
                    if val is not None and not pd.isna(val):
                        pm_sched[yr] = float(val)
                except Exception:
                    pass

    return fitout_breakdown, capex_sched, pm_sched
