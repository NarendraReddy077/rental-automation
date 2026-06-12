import datetime
import io
import pandas as pd
import openpyxl
from sheets.capex_pm import calculate_fy_months

def parse_input_xlsx(file_bytes):
    """Parses user-uploaded input Excel sheet with extensive error resilience."""
    try:
        # Resolve raw bytes regardless of input type (bytes or file-like)
        if hasattr(file_bytes, "getvalue"):
            raw_data = file_bytes.getvalue()
        elif hasattr(file_bytes, "read"):
            try:
                file_bytes.seek(0)
            except Exception:
                pass
            raw_data = file_bytes.read()
        else:
            raw_data = file_bytes
            
        df = pd.read_excel(io.BytesIO(raw_data), sheet_name="Main")
        df.columns = [str(c).strip() for c in df.columns]
        
        # Dynamically determine Field Name and Sample Value columns
        field_col = None
        value_col = None
        for col in df.columns:
            col_lower = col.lower()
            if "field" in col_lower or "parameter" in col_lower or "name" in col_lower:
                field_col = col
            elif "value" in col_lower or "sample" in col_lower:
                value_col = col
        
        if field_col is None:
            field_col = df.columns[0] if len(df.columns) > 0 else "Field Name"
        if value_col is None:
            value_col = df.columns[1] if len(df.columns) > 1 else "Sample Value"

        # Build dictionary from Name-Value pairs
        raw_params = {}
        current_section = "rent"
        if field_col in df.columns and value_col in df.columns:
            for idx, row in df.iterrows():
                name = str(row[field_col]).strip()
                val = row[value_col]
                if pd.isna(val):
                    val = None
                
                # Update current section based on keywords in name
                name_lower = name.lower()
                if "cam" in name_lower:
                    current_section = "cam"
                elif "parking" in name_lower or "wheeler" in name_lower:
                    current_section = "parking"
                elif "fitout" in name_lower or "capex" in name_lower:
                    current_section = "capex"
                elif "pm" in name_lower or "maintenance" in name_lower:
                    current_section = "pm"
                
                # If the key is a generic escalation key, qualify it with the section name
                key_to_store = name
                if name_lower in ["escalation %", "escalation frequency months", "escalation frequency"]:
                    if current_section == "cam":
                        key_to_store = f"CAM {name}"
                    elif current_section == "parking":
                        key_to_store = f"Parking {name}"
                    elif current_section == "rent":
                        key_to_store = f"Rent {name}"
                
                raw_params[key_to_store] = val
        
        # Helper to parse dates
        def parse_date(date_val, default=None):
            if date_val is None or pd.isna(date_val):
                return default
            if isinstance(date_val, (datetime.date, datetime.datetime)):
                return date_val.date() if isinstance(date_val, datetime.datetime) else date_val
            try:
                return pd.to_datetime(str(date_val).strip()).date()
            except Exception:
                return default

        # Helper to check key and all aliases (case-insensitive and stripped)
        def get_value_by_aliases(aliases, default=None):
            for alias in aliases:
                alias_clean = alias.lower().strip()
                for key, val in raw_params.items():
                    if key.lower().strip() == alias_clean:
                        return val
            return default

        # Safely extract values from raw_params with fallbacks and aliases
        def get_str(name, aliases, default):
            v = get_value_by_aliases([name] + aliases)
            if v is None or pd.isna(v) or str(v).strip() == "":
                return default
            return str(v).strip()

        def get_float(name, aliases, default):
            v = get_value_by_aliases([name] + aliases)
            if v is None or pd.isna(v):
                return default
            try:
                return float(v)
            except Exception:
                return default

        def get_int(name, aliases, default):
            v = get_value_by_aliases([name] + aliases)
            if v is None or pd.isna(v):
                return default
            try:
                return int(float(v))
            except Exception:
                return default

        # Normalize values
        params = {}
        params["Lease ID"] = get_str("Lease ID", ["LeaseID", "Lease_ID"], "LEASE-001")
        params["REU Name"] = get_str("REU Name", ["REUName", "REU_Name", "REU"], "IN-CHEK2")
        params["Lease Type"] = get_str("Lease Type", ["LeaseType", "Lease_Type"], "Office")
        params["Building Name"] = get_str("Building Name", ["BuildingName", "Building_Name"], "SKCL Tech Park")
        params["City"] = get_str("City", [], "Chennai")
        params["Country"] = get_str("Country", [], "India")
        params["Currency"] = get_str("Currency", [], "INR")
        
        params["Chargeable Area Sqft"] = get_float("Chargeable Area Sqft", ["Chargeable Area Sq.ft.", "Chargeable Area Sq.ft", "Chargeable Area (Sqft)", "Chargeable Area", "Area Sqft", "Area", "Chargeable Office Area"], 9515.0)
        params["Parking Slots"] = get_int("Parking Slots", ["ParkingSlots", "Parking Slots Count"], 20)
        params["4 Wheeler Slots"] = get_int("4 Wheeler Slots", ["No. of 4 wheelers", "4-wheeler slots", "4 Wheeler Parking Slots", "No of 4 wheelers"], 80)
        params["4 Wheeler Rate"] = get_float("4 Wheeler Rate", ["Rate per 4 wheeler", "4-wheeler rate", "4 Wheeler Parking Rate"], 1500.0)
        params["2 Wheeler Slots"] = get_int("2 Wheeler Slots", ["No. of 2 wheelers", "2-wheeler slots", "2 Wheeler Parking Slots", "No of 2 wheelers"], 50)
        params["2 Wheeler Rate"] = get_float("2 Wheeler Rate", ["Rate per 2 wheeler", "2-wheeler rate", "2 Wheeler Parking Rate"], 1000.0)
        
        start_date_val = get_value_by_aliases(["Agreement Start Date", "Start Date", "Lease Start Date", "Commencement Date"])
        start_date = parse_date(start_date_val, datetime.date(2026, 4, 1))
        
        end_date_val = get_value_by_aliases(["Agreement End Date", "End Date", "Lease End Date", "Expiry Date"])
        end_date = parse_date(end_date_val, datetime.date(2031, 3, 31))
        
        params["Agreement Start Date"] = start_date
        params["Agreement End Date"] = end_date
        
        rent_start_date_val = get_value_by_aliases(["Rent Start Date", "Rent Commencement Date"])
        params["Rent Start Date"] = parse_date(rent_start_date_val, start_date)
        
        # Lease Term calculations
        term_months = get_value_by_aliases(["Lease Term Months", "Lease Term", "Term Months", "Duration Months"])
        if term_months is None or pd.isna(term_months):
            # Calculate months from start/end dates
            term_months = round((end_date - start_date).days / 30.4167)
        else:
            try:
                term_months = int(float(term_months))
            except Exception:
                term_months = round((end_date - start_date).days / 30.4167)
        params["Lease Term Months"] = term_months
        
        params["Rent Per Sqft"] = get_float("Rent Per Sqft", ["Rent Per Sq.ft.", "Rent per Sqft", "Rent Per Sqft/month", "Base Rent", "Quoted Rentals", "Quoted Rentals "], 120.0)
        params["Quoted CAM"] = get_float("Quoted CAM", ["CAM", "CAM Per Sqft", "Quoted CAM (per sq ft/month)", "CAM per Sq ft", "CAM per Sq.ft.", "CAM per Sqft", "CAM Rate"], 15.48)
        
        esc_val = get_value_by_aliases(["Escalation %", "Rent Escalation %", "Escalation Percentage", "Rent Escalation Pct"])
        params["Escalation %"] = float(esc_val) if esc_val is not None and not pd.isna(esc_val) else 0.15
        
        # Robust check for frequency e.g. 0.36 -> 36 months
        freq_val = get_value_by_aliases(["Escalation Frequency Months", "Rent Escalation Frequency Months", "Rent Escalation Frequency", "Escalation Freq"])
        if freq_val is not None and not pd.isna(freq_val):
            try:
                freq_val = float(freq_val)
                if freq_val < 1.0:
                    freq_val = int(round(freq_val * 100))
                else:
                    freq_val = int(freq_val)
            except Exception:
                freq_val = 36
        else:
            freq_val = 36
        params["Escalation Frequency Months"] = freq_val
        
        cam_esc_val = get_value_by_aliases(["CAM Escalation %", "CAM Escalation Percentage", "CAM Escalation Pct", "CAM Escalation", "CAM escalation %"])
        params["CAM Escalation %"] = float(cam_esc_val) if cam_esc_val is not None and not pd.isna(cam_esc_val) else 0.05
        
        cam_freq_val = get_value_by_aliases(["CAM Escalation Frequency Months", "CAM Escalation Frequency", "CAM Escalation Freq", "CAM Escalation Period", "CAM Escalation Period Months", "escalation period", "CAM escalation period", "CAM escalation frequency"])
        if cam_freq_val is not None and not pd.isna(cam_freq_val):
            try:
                cam_freq_val = float(cam_freq_val)
                if cam_freq_val < 1.0:
                    cam_freq_val = int(round(cam_freq_val * 100))
                else:
                    cam_freq_val = int(cam_freq_val)
            except Exception:
                cam_freq_val = 12
        else:
            cam_freq_val = 12
        params["CAM Escalation Frequency Months"] = cam_freq_val
        
        # Parking escalation parameters (fallback to Rent Escalation if not specified)
        park_esc_val = get_value_by_aliases(["Parking Escalation %", "Parking Escalation Percentage", "Parking Escalation Pct"])
        if park_esc_val is not None and not pd.isna(park_esc_val):
            params["Parking Escalation %"] = float(park_esc_val)
        else:
            params["Parking Escalation %"] = params["Escalation %"]
            
        park_freq_val = get_value_by_aliases(["Parking Escalation Frequency Months", "Parking Escalation Frequency", "Parking Escalation Freq"])
        if park_freq_val is not None and not pd.isna(park_freq_val):
            try:
                park_freq_val = float(park_freq_val)
                if park_freq_val < 1.0:
                    park_freq_val = int(round(park_freq_val * 100))
                else:
                    park_freq_val = int(park_freq_val)
            except Exception:
                park_freq_val = params["Escalation Frequency Months"]
        else:
            park_freq_val = params["Escalation Frequency Months"]
        params["Parking Escalation Frequency Months"] = park_freq_val

        params["Billing Frequency"] = get_str("Billing Frequency", ["BillingFrequency"], "Monthly")
        params["Security Deposit Months"] = get_float("Security Deposit Months", ["Security Deposit (number of months)", "Security Deposit (Months)"], 6.0)
        params["Security Deposit Amount"] = get_float("Security Deposit Amount", ["Security Deposit Amount (INR)", "Security Deposit"], 11418000.0)
        params["Refundable Deposit"] = get_str("Refundable Deposit", ["Refundable"], "Yes")
        
        # Capex & Cost factors
        total_fitout = get_float("Fitout Cost", ["Total Fitout Cost", "Fitout Cost Amount", "Fitouts", "Capex"], 34000000.0)
        params["Fitout Cost"] = total_fitout
        params["Useful Life Years"] = get_float("Useful Life Years", ["Useful Life", "Useful Life (Years)"], 5.0)
        params["Residual Value"] = get_float("Residual Value", ["ResidualValue"], 0.0)
        params["Discount Rate"] = get_float("Discount Rate", ["Discounting Rate", "Discount Rate %"], 0.08)
        params["Incremental Borrowing Rate"] = get_float("Incremental Borrowing Rate", ["IBR", "Incremental Borrowing Rate %", "IBR %"], 0.08)
        params["Cost of Capital"] = get_float("Cost of Capital", ["WACC", "Cost of Capital (WACC) %", "WACC %"], 0.105)
        params["Addnl.Deposit -energy(Refundable)"] = get_float("Addnl.Deposit -energy(Refundable)", ["Energy Deposit", "Energy Security Deposit", "Energy Deposit Amount"], 500000.0)
        
        # Add new parameters missing in original but parsed if they exist
        params["Imputed Interest Rate"] = get_float("Imputed Interest Rate", ["Imputed Interest Rate %", "Imputed Interest", "Imputed Rate"], 0.0711)
        params["Ready Reckoner Rate"] = get_float("Ready Reckoner Rate", ["Ready Reckoner Rate (INR/sq m)", "Ready Reckoner"], 15000.0)
        params["Exchange Rate"] = get_float("Exchange Rate", ["Forex Rate", "Exchange Rate (INR/Euro)", "Forex"], 105.02)
        params["Incremental Restoration Cost Sqft"] = get_float("Incremental Restoration Cost Sqft", ["Restoration Cost per Sq ft (ARO)", "Restoration Cost per Sqft", "ARO Rate", "ARO Cost per Sqft"], None)
        
        # OpEx I and OpEx II parameters
        params["Opex Others Per Month"] = get_float("Opex Others Per Month", ["Opex Others", "OpEx Others Rs../ month (At Actuals)", "Opex Others Rs./ month", "Opex Others Rs/ month", "Opex I"], 654.0)
        params["Opex II Per Month"] = get_float("Opex II Per Month", ["Opex II", "OpEx II Rs../ month", "Opex II Rs./ month", "Opex II Rs/ month", "Opex II - OpEx Add-on"], 0.0)
 
        
        total_pm = get_float("PM Cost Over Lease", ["Preventive Maintenance Cost", "PM Cost", "Maintenance Cost over Lease", "PM cost Over lease period", "PM cost Over lease"], 2500000.0)
        params["PM Cost Over Lease"] = total_pm

        # Try to parse dynamic schedules from CAPEX and PM sheet if it exists
        fitout_breakdown = []
        fitout_lifes = []
        fitout_active_months_breakdown = []
        capex_sched = {}
        capex_lives = {}
        pm_sched = {}
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
            if "Lease Rent" in wb.sheetnames:
                ws_lr = wb["Lease Rent"]
                val_c22 = ws_lr["C22"].value
                val_d22 = ws_lr["D22"].value
                val_c23 = ws_lr["C23"].value
                val_d23 = ws_lr["D23"].value
                if val_c22 is not None:
                    try:
                        params["4 Wheeler Rate"] = float(val_c22)
                    except Exception:
                        pass
                if val_d22 is not None:
                    try:
                        params["4 Wheeler Slots"] = int(float(val_d22))
                    except Exception:
                        pass
                if val_c23 is not None:
                    try:
                        params["2 Wheeler Rate"] = float(val_c23)
                    except Exception:
                        pass
                if val_d23 is not None:
                    try:
                        params["2 Wheeler Slots"] = int(float(val_d23))
                    except Exception:
                        pass
            if "Rent Calculation" in wb.sheetnames:
                ws_rc = wb["Rent Calculation"]
                val_j5 = ws_rc["J5"].value
                if val_j5 is not None:
                    try:
                        params["Exchange Rate"] = float(val_j5)
                    except Exception:
                        pass
            if "CAPEX and PM" in wb.sheetnames:
                ws = wb["CAPEX and PM"]
                
                # 1. Parse fitout costs, useful lives and active months dynamically (up to 8 phases)
                for k in range(8):
                    r_c = 6 + 6 * k
                    r_m = 3 + 6 * k
                    # Check if this phase exists in sheet (we check row label at column A)
                    lbl = ws.cell(row=r_c, column=1).value
                    if lbl == "Investment Cost":
                        cost_val = ws.cell(row=r_c, column=2).value
                        life_val_months = ws.cell(row=r_m, column=5).value
                        
                        cost = float(cost_val) if isinstance(cost_val, (int, float)) else 0.0
                        life = int(life_val_months) if isinstance(life_val_months, (int, float)) else 72
                        
                        fitout_breakdown.append(cost)
                        fitout_lifes.append(life)
                        
                        # Parse active months for this phase year-by-year (columns F to O)
                        phase_m = {}
                        for c in range(6, 16):
                            yr_val = ws.cell(row=2, column=c).value
                            m_val = ws.cell(row=r_m, column=c).value
                            try:
                                yr = int(float(str(yr_val).strip()))
                            except (ValueError, TypeError):
                                continue
                            phase_m[yr] = int(m_val) if isinstance(m_val, (int, float)) else 0
                        fitout_active_months_breakdown.append(phase_m)
                
                # 2. Parse Capex schedule and useful lives
                r_capex_header = 36
                for r in range(1, 100):
                    if ws.cell(row=r, column=1).value == "Investment name" and ws.cell(row=r, column=2).value == "Capex":
                        r_capex_header = r
                        break
                
                # Capex values are in row r_capex_header + 2
                r_capex_val = r_capex_header + 2
                for c in range(6, 16):
                    yr_val = ws.cell(row=r_capex_header, column=c).value
                    val = ws.cell(row=r_capex_val, column=c).value
                    # Tranche i active months row is: r_capex_header + 4 + 5 * (i-1)
                    tranche_row = r_capex_header + 4 + 5 * (c - 6)
                    life_val = ws.cell(row=tranche_row, column=5).value
                    try:
                        yr = int(float(str(yr_val).strip()))
                    except (ValueError, TypeError):
                        continue
                    try:
                        val_num = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        val_num = 0.0
                    if val_num > 0:
                        capex_sched[yr] = val_num
                        try:
                            capex_lives[yr] = int(life_val) if life_val is not None else 0
                        except (ValueError, TypeError):
                            pass
                            
                # 3. Parse PM schedule from PM row. Let's find PM row dynamically!
                # PM is the row where column A is "Basic and project maintenance"
                r_pm = 95
                for r in range(1, 150):
                    if ws.cell(row=r, column=1).value == "Basic and project maintenance":
                        r_pm = r
                        break
                # PM year headers are in row 2 (which is always row 2)
                for c in range(6, 16):
                    yr_val = ws.cell(row=2, column=c).value
                    val = ws.cell(row=r_pm, column=c).value
                    try:
                        yr = int(float(str(yr_val).strip()))
                    except (ValueError, TypeError):
                        continue
                    try:
                        val_num = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        val_num = 0.0
                    if val_num > 0:
                        pm_sched[yr] = val_num
            else:
                # Parse from Main sheet Name-Value pairs (raw_params)
                # 1. Fitout breakdown costs
                fitout_costs_dict = {}
                for key, val in raw_params.items():
                    key_clean = key.lower()
                    if "fitout" in key_clean and "phase" in key_clean:
                        import re
                        m = re.search(r"\d+", key)
                        if m:
                            try:
                                phase_num = int(m.group(0))
                                if val is not None and not pd.isna(val):
                                    fitout_costs_dict[phase_num] = float(val)
                            except Exception:
                                pass
                for p_num in sorted(fitout_costs_dict.keys()):
                    fitout_breakdown.append(fitout_costs_dict[p_num])
                            
                # 2. Capex schedule
                for key, val in raw_params.items():
                    key_clean = key.lower().replace(" ", "")
                    if "capex" in key_clean:
                        import re
                        m = re.search(r"\b(20\d{2})\b|(?<=fy)(20\d{2})", key_clean)
                        if not m:
                            m = re.search(r"\d{4}", key)
                        if m:
                            try:
                                yr = int(m.group(0))
                                if val is not None and not pd.isna(val):
                                    capex_sched[yr] = float(val)
                            except Exception:
                                pass
                            
                # 3. PM schedule
                for key, val in raw_params.items():
                    key_clean = key.lower().replace(" ", "")
                    if "pm" in key_clean or "maintenance" in key_clean:
                        import re
                        m = re.search(r"\b(20\d{2})\b|(?<=fy)(20\d{2})", key_clean)
                        if not m:
                            m = re.search(r"\d{4}", key)
                        if m:
                            try:
                                yr = int(m.group(0))
                                if val is not None and not pd.isna(val):
                                    pm_sched[yr] = float(val)
                            except Exception:
                                pass
        except Exception:
            pass

        # Fallbacks for empty/non-existent sheet cases
        start_year = start_date.year
        term_months = params["Lease Term Months"]
        
        was_breakdown_parsed = (len(fitout_breakdown) > 0)
        
        # If fitout breakdown was not parsed, initialize default 3 phases
        if not fitout_breakdown:
            fitout_breakdown = [14000000.0, 5000000.0, 15000000.0]
            fitout_lifes = [72, 30, 48]
            
        if was_breakdown_parsed:
            # Update Fitout Cost to match sum of parsed breakdown
            total_fitout = sum(fitout_breakdown)
            params["Fitout Cost"] = total_fitout
        else:
            # If fitout total from Main doesn't match sum of parsed breakdown, scale it
            if abs(total_fitout - sum(fitout_breakdown)) > 1.0:
                if sum(fitout_breakdown) > 0:
                    scale = total_fitout / sum(fitout_breakdown)
                    fitout_breakdown = [f * scale for f in fitout_breakdown]
                else:
                    num_p = len(fitout_breakdown) if fitout_breakdown else 3
                    fitout_breakdown = [total_fitout / num_p] * num_p

        # If fitout_lifes is empty, initialize default lives
        if not fitout_lifes:
            default_lifes = [72, 30, 48]
            fitout_lifes = []
            for idx in range(len(fitout_breakdown)):
                if idx < len(default_lifes):
                    fitout_lifes.append(default_lifes[idx])
                else:
                    fitout_lifes.append(int(params["Useful Life Years"] * 12))

        # Fitout active months fallback
        if not fitout_active_months_breakdown:
            fitout_active_months_breakdown = []
            for life in fitout_lifes:
                fitout_active_months_breakdown.append(
                    calculate_fy_months(start_date, max(term_months, 120), life)
                )

        # Capex schedule fallback
        if not capex_sched:
            capex_sched = {
                start_year: 14000000.0,
                start_year + 1: 12000000.0,
                start_year + 2: 5000000.0,
                start_year + 3: 6000000.0,
                start_year + 4: 6000000.0
            }
            # Scale Capex if total Fitout Cost changed
            if abs(total_fitout - 34000000.0) > 1.0:
                scale = total_fitout / 34000000.0
                for k in capex_sched:
                    capex_sched[k] *= scale

        # Capex useful lives fallback
        if not capex_lives:
            for idx, yr in enumerate(sorted(capex_sched.keys())):
                if idx == 0:
                    life = term_months
                else:
                    first_year_months = 12 - start_date.month + 1
                    elapsed = first_year_months + 12 * (idx - 1)
                    life = max(0, term_months - elapsed)
                capex_lives[yr] = life

        # PM schedule fallback
        was_pm_parsed = (len(pm_sched) > 0)
        if not pm_sched:
            pm_sched = {
                start_year: 1500000.0,
                start_year + 1: 200000.0,
                start_year + 2: 200000.0,
                start_year + 3: 200000.0,
                start_year + 4: 200000.0,
                start_year + 5: 200000.0
            }
            # Scale PM if total PM cost changed
            if abs(total_pm - 2500000.0) > 1.0:
                scale = total_pm / 2500000.0
                for k in pm_sched:
                    pm_sched[k] *= scale
                    
        if was_pm_parsed:
            total_pm = sum(pm_sched.values())
            params["PM Cost Over Lease"] = total_pm
                    
        params["Fitout Cost Breakdown"] = fitout_breakdown
        params["Fitout Useful Lives"] = fitout_lifes
        params["Fitout Active Months Breakdown"] = fitout_active_months_breakdown
        params["Capex Schedule"] = capex_sched
        params["Capex Useful Lives"] = capex_lives
        params["PM Schedule"] = pm_sched

        return params, None
    except Exception as e:
        return {}, str(e)
