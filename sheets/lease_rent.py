import openpyxl
import math
import datetime

def get_edate_minus_1(d, m_offset):
    y = d.year + (d.month + m_offset - 1) // 12
    m = (d.month + m_offset - 1) % 12 + 1
    import calendar
    last_day = calendar.monthrange(y, m)[1]
    day = min(d.day, last_day)
    return datetime.date(y, m, day) - datetime.timedelta(days=1)

def inject(ws, params):
    """
    Dynamically injects rental and CAM escalation formulas/values into 'Lease Rent' worksheet.
    Also inserts rows if term_months > 72, and updates cross-sheet references in 'Rent Calculation'.
    """
    rent_rate = params.get("Rent Per Sqft", 120.0)
    cam_rate = params.get("Quoted CAM", 15.48)
    rent_esc_pct = params.get("Escalation %", 0.15)
    rent_esc_freq = params.get("Escalation Frequency Months", 36)
    cam_esc_pct = params.get("CAM Escalation %", 0.05)
    cam_esc_freq = params.get("CAM Escalation Frequency Months", 12)
    parking_esc_pct = params.get("Parking Escalation %", rent_esc_pct)
    parking_esc_freq = params.get("Parking Escalation Frequency Months", rent_esc_freq)
    term_months = params.get("Lease Term Months", 72)
    
    # Calculate y1_months dynamically based on start_date and October-September financial year
    start_date = params.get("Agreement Start Date")
    if isinstance(start_date, str):
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date = None
            
    if start_date is None:
        start_date = datetime.date(2026, 4, 1)
        
    if start_date.month <= 9:
        fy_end_year = start_date.year
    else:
        fy_end_year = start_date.year + 1
        
    first_fy_end = datetime.date(fy_end_year, 9, 30)
        
    y1_months = 0
    for m in range(1, 100):
        end_d = get_edate_minus_1(start_date, m)
        if end_d <= first_fy_end:
            y1_months = m
        else:
            break
    y1_months = max(1, y1_months)

    # Calculate how many months we actually need to write
    # If the user-defined lease is longer than 72 months, we dynamically insert rows.
    # The default specimen template has 72 monthly rows (ending at row 102).
    # Month 1 is row 31. Month 72 is row 102.
    # If term_months > 72, we insert (term_months - 72) rows at Row 103 (above original summary row).
    max_months = term_months
    if term_months > 72:
        inserted = term_months - 72
        ws.insert_rows(103, amount=inserted)
    elif term_months < 72:
        deleted = 72 - term_months
        ws.delete_rows(31 + term_months, amount=deleted)

    ws["C22"] = params.get("4 Wheeler Rate", 1500.0)
    ws["D22"] = params.get("4 Wheeler Slots", 80)
    ws["C23"] = params.get("2 Wheeler Rate", 1000.0)
    ws["D23"] = params.get("2 Wheeler Slots", 50)

    # Update starting date and calendar year dynamically
    ws["B31"] = "=Main!B4"
    ws["A31"] = "=YEAR(B31)"
    ws["G19"] = "=YEAR(B31)"

    # Identify the year boundary rows
    summary_rows = {}
    
    # Loop through all months and write formulas
    for m in range(1, max_months + 1):
        r = 30 + m
        
        # Beyond the lease term, set rates to 0.0
        if m > term_months:
            ws.cell(row=r, column=1, value="")   # Column A: Year
            ws.cell(row=r, column=2, value="")   # Column B: From Date
            ws.cell(row=r, column=3, value="")   # Column C: To Date
            ws.cell(row=r, column=4, value="")   # Column D: Months
            ws.cell(row=r, column=5, value="")   # Column E: Area SQF
            ws.cell(row=r, column=6, value="")   # Column F: Area SQM
            ws.cell(row=r, column=7, value="")   # Column G: Escalation
            ws.cell(row=r, column=8, value=0.0)  # Column H: Rental rate
            ws.cell(row=r, column=9, value="")   # Column I: Car Parks
            ws.cell(row=r, column=10, value=0.0) # Column J: CAM rate
            ws.cell(row=r, column=11, value=0.0) # Column K: Rental SQF
            ws.cell(row=r, column=12, value=0.0) # Column L: CAM SQF
            ws.cell(row=r, column=13, value=0.0) # Column M: Total
            continue
            
        # Write Date, Month, Escalation, Rate, and Amount columns
        if m == 1:
            ws.cell(row=r, column=1, value="=YEAR(B31)")
            ws.cell(row=r, column=2, value="=Main!B4")
            ws.cell(row=r, column=5, value="='Rent Calculation'!$F$23")
            ws.cell(row=r, column=6, value="='Rent Calculation'!G23")
            ws.cell(row=r, column=8, value="=Main!B7")
            ws.cell(row=r, column=9, value="=E24")
            ws.cell(row=r, column=10, value="=Main!B8")
        else:
            # Year Column A
            if m > y1_months and (m - y1_months - 1) % 12 == 0:
                ws.cell(row=r, column=1, value=f"=A{r-1}+1")
            else:
                ws.cell(row=r, column=1, value=f"=A{r-1}")
            
            # From Date Column B
            ws.cell(row=r, column=2, value=f"=C{r-1}+1")
            
            ws.cell(row=r, column=5, value=f"=E{r-1}")
            ws.cell(row=r, column=6, value=f"=F{r-1}")
            
            # Rent Escalation
            if rent_esc_freq and rent_esc_freq > 0 and (m - 1) % rent_esc_freq == 0:
                ws.cell(row=r, column=8, value=f"=H{r-1}*(1+{rent_esc_pct})")
            else:
                ws.cell(row=r, column=8, value=f"=H{r-1}")
                
            # Parking Escalation
            if parking_esc_freq and parking_esc_freq > 0 and (m - 1) % parking_esc_freq == 0:
                ws.cell(row=r, column=9, value=f"=I{r-1}*(1+{parking_esc_pct})")
            else:
                ws.cell(row=r, column=9, value=f"=I{r-1}")
                
            # CAM Escalation
            if cam_esc_freq and cam_esc_freq > 0 and (m - 1) % cam_esc_freq == 0:
                ws.cell(row=r, column=10, value=f"=J{r-1}*(1+{cam_esc_pct})")
            else:
                ws.cell(row=r, column=10, value=f"=J{r-1}")
                
        # Common columns for all active months
        ws.cell(row=r, column=3, value=f"=EDATE(B{r},1)-1") # Column C: To Date
        ws.cell(row=r, column=4, value=1)                  # Column D: Months
        ws.cell(row=r, column=7, value="")                 # Column G: Escalation
        ws.cell(row=r, column=11, value=f"=H{r}*E{r}")     # Column K: Rental SQF
        ws.cell(row=r, column=12, value=f"=J{r}*E{r}")     # Column L: CAM SQF
        ws.cell(row=r, column=13, value=f"=I{r}+K{r}+L{r}") # Column M: Total
                
        # Determine if this row is a Year boundary (for years 1 to 10)
        # Year 1 summary is written at row 39 (Month 9)
        # Year 2 summary at row 51 (Month 21)
        # etc.
        # So boundary row is when m = 8 + 12 * (y - 1) + 1 = 9 + 12 * (y - 1)
        if m > y1_months and (m - y1_months - 1) % 12 == 0:
            y = (m - y1_months - 1) // 12 + 1  # This is the year that just ended
            if y <= 10:
                if y == 1:
                    start_sum = 31
                    end_sum = 30 + y1_months
                else:
                    start_sum = 31 + y1_months + 12 * (y - 2)
                    end_sum = start_sum + 11
                ws.cell(row=r, column=14, value=f"=SUM(K{start_sum}:K{end_sum})") # Column N: Rentals
                ws.cell(row=r, column=15, value=f"=SUM(L{start_sum}:L{end_sum})") # Column O: CAM
                ws.cell(row=r, column=16, value=f"=SUM(I{start_sum}:I{end_sum})") # Column P: Parking
                summary_rows[y] = r

    # Calculate final year number and the final summary row index
    if term_months <= y1_months:
        final_year = 1
        start_sum = 31
        end_sum = 30 + term_months
    else:
        final_year = math.ceil((term_months - y1_months) / 12) + 1
        if final_year == 1:
            start_sum = 31
        else:
            start_sum = 31 + y1_months + 12 * (final_year - 2)
        end_sum = 30 + term_months
        
    final_sum_row = 30 + term_months + 1
    
    # Write the final year summary row
    ws.cell(row=final_sum_row, column=1, value="")
    ws.cell(row=final_sum_row, column=2, value="")
    ws.cell(row=final_sum_row, column=14, value=f"=SUM(K{start_sum}:K{end_sum})") # Column N
    ws.cell(row=final_sum_row, column=15, value=f"=SUM(L{start_sum}:L{end_sum})") # Column O
    ws.cell(row=final_sum_row, column=16, value=f"=SUM(I{start_sum}:I{end_sum})") # Column P
    summary_rows[final_year] = final_sum_row
    
    # Write the grand total row at final_sum_row + 1
    total_row = final_sum_row + 1
    ws.cell(row=total_row, column=9, value=f"=SUM(I31:I{30 + term_months})") # Column I
    ws.cell(row=total_row, column=11, value=f"=SUM(K31:K{30 + term_months})") # Column K
    ws.cell(row=total_row, column=12, value=f"=SUM(L31:L{30 + term_months})") # Column L
    ws.cell(row=total_row, column=13, value=f"=SUM(M31:M{30 + term_months})") # Column M
    ws.cell(row=total_row, column=14, value=f"=SUM(N31:N{final_sum_row})") # Column N
    ws.cell(row=total_row, column=15, value=f"=SUM(O31:O{final_sum_row})") # Column O
    ws.cell(row=total_row, column=16, value=f"=SUM(P31:P{final_sum_row})") # Column P

    # Update calendar year summary table (rows 19 to 25) to reference correct monthly row bounds
    for r_sum in range(19, 26):
        ws.cell(row=r_sum, column=8, value=f"=SUMIF($A$31:$A$1048576,$G{r_sum},$D$31:$D{30 + term_months})")
        ws.cell(row=r_sum, column=9, value=f"=SUMIF($A$31:$A$1048576,$G{r_sum},$K$31:$K{30 + term_months})")
        ws.cell(row=r_sum, column=10, value=f"=SUMIF($A$31:$A$1048576,$G{r_sum},$L$31:$L{30 + term_months})")

    # Update Rent Calculation sheet references dynamically using ws.parent
    wb = ws.parent
    if "Rent Calculation" in wb.sheetnames:
        rc_ws = wb["Rent Calculation"]
        
        # Years 1 to 10 correspond to Rent Calculation rows 210 to 219
        for y in range(1, 11):
            r_rc = 209 + y
            if y in summary_rows:
                sum_row = summary_rows[y]
                rc_ws.cell(row=r_rc, column=3, value=f"='Lease Rent'!N{sum_row}") # Column C: Rent (INR)
                rc_ws.cell(row=r_rc, column=7, value=f"='Lease Rent'!P{sum_row}") # Column G: Car Park (INR)
                rc_ws.cell(row=r_rc, column=9, value=f"='Lease Rent'!O{sum_row}") # Column I: CAM (INR)
            else:
                # Inactive year: set to 0.0
                rc_ws.cell(row=r_rc, column=3, value=0.0)
                rc_ws.cell(row=r_rc, column=7, value=0.0)
                rc_ws.cell(row=r_rc, column=9, value=0.0)
