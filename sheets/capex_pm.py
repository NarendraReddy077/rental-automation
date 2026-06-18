import pandas as pd
import numpy as np
import datetime
import openpyxl

def get_y1_months(start_date):
    """
    Calculates the number of months in the first financial year (Oct-Sep)
    from start_date. Matches the lease_rent.py logic.
    """
    if start_date.month <= 9:
        fy_end_year = start_date.year
    else:
        fy_end_year = start_date.year + 1
        
    first_fy_end = datetime.date(fy_end_year, 9, 30)
    
    y1_months = 0
    for m in range(1, 13):
        y = start_date.year + (start_date.month + m - 1) // 12
        m_calc = (start_date.month + m - 1) % 12 + 1
        import calendar
        last_day = calendar.monthrange(y, m_calc)[1]
        day = min(start_date.day, last_day)
        end_d = datetime.date(y, m_calc, day) - datetime.timedelta(days=1)
        
        if end_d <= first_fy_end:
            y1_months = m
        else:
            break
    return max(1, y1_months)

def calculate_fy_months(start_date, term_months, fitout_term_months):
    """
    Dynamically distributes term_months across fiscal years (Oct to Sep).
    Keys are calendar years representing the financial year of the lease.
    """
    fy_months = {}
    remaining_months = min(term_months, fitout_term_months)
    if remaining_months <= 0:
        return fy_months
        
    y1_months = get_y1_months(start_date)
    
    current_key_year = start_date.year
    first_year_m = min(remaining_months, y1_months)
    fy_months[current_key_year] = first_year_m
    remaining_months -= first_year_m
    
    # Subsequent years
    while remaining_months > 0:
        current_key_year += 1
        months = min(remaining_months, 12)
        fy_months[current_key_year] = months
        remaining_months -= months
        
    return fy_months

def get_capex_tranche_months(start_date, term_months, tranche_idx, useful_life):
    """
    Helper to calculate active months for a given CAPEX tranche with user-defined useful life.
    Capex is allowed to extend up to 120 months (10 years) like Fitouts.
    """
    effective_term = max(term_months, 120)
    if tranche_idx == 1:
        return calculate_fy_months(start_date, effective_term, useful_life)
        
    y1_months = get_y1_months(start_date)
    
    elapsed = y1_months + 12 * (tranche_idx - 2)
    remaining = max(0, effective_term - elapsed)
    if remaining <= 0:
        return {}
        
    amort_term = min(remaining, useful_life)
    fy_months = {}
    current_key_year = start_date.year + tranche_idx - 1
    
    first_year_m = min(amort_term, y1_months)
    fy_months[current_key_year] = first_year_m
    amort_term -= first_year_m
    
    current_key_year += 1
    while amort_term > 0:
        months = min(amort_term, 12)
        fy_months[current_key_year] = months
        amort_term -= months
        current_key_year += 1
        
    return fy_months

def shift_merged_ranges(ws, start_row, amount):
    """
    Shifts merged cell ranges starting at or after start_row by amount.
    """
    ranges = list(ws.merged_cells.ranges)
    for r in ranges:
        if r.min_row >= start_row:
            ws.merged_cells.remove(r)
            r.shift(row_shift=amount)
            ws.merged_cells.add(r)

def inject(ws, params):
    """
    Injects parameters and schedules into 'CAPEX and PM' worksheet.
    Handles dynamic insertion of new Fitout rows if N > 3.
    """
    start_date = params["Agreement Start Date"]
    start_year = start_date.year
    term_months = params.get("Lease Term Months", 72)
    if pd.isna(term_months) or term_months is None:
        term_months = 72

    # Write dynamic relative year headers in Row 2
    for c in range(6, 16):
        if c == 6:
            ws.cell(row=2, column=c, value="=YEAR('Lease Rent'!B31)")
        else:
            prev_col = openpyxl.utils.get_column_letter(c - 1)
            ws.cell(row=2, column=c, value=f"={prev_col}2+1")

    fitouts = list(params.get("Fitout Cost Breakdown", []))
    fitout_lifes = list(params.get("Fitout Useful Lives", []))
    fitout_active_months = params.get("Fitout Active Months Breakdown", [])
    
    N = len(fitouts)
    offset = 6 * (N - 3) if N > 3 else 0
    
    # Insert new rows above row 20 if N > 3
    if N > 3:
        ws.insert_rows(20, amount=offset)
        shift_merged_ranges(ws, 20, offset)
        
    # Write each Fitout phase
    for k in range(N):
        r_m = 3 + 6 * k      # Month needed row
        r_c = 6 + 6 * k      # Investment Cost row
        r_d = 7 + 6 * k      # Depreciation Time row
        
        ws.cell(row=r_m, column=1, value=f"Fitout Phase {k+1}")
        ws.cell(row=r_m, column=4, value="Month needed")
        ws.cell(row=r_m, column=5, value=f"=SUM(F{r_m}:O{r_m})")
        
        ws.cell(row=r_m+1, column=4, value="Month left")
        ws.cell(row=r_m+1, column=5, value=f"=E{r_m}-SUM(F{r_m}:O{r_m})")
        
        ws.cell(row=r_c, column=1, value="Investment Cost")
        ws.cell(row=r_c, column=2, value=fitouts[k])
        ws.cell(row=r_c, column=4, value="Annual Depreciation Rate")
        
        ws.cell(row=r_d, column=1, value="Depreciation Time")
        ws.cell(row=r_d, column=2, value=f"=E{r_m}/12")
        ws.cell(row=r_d, column=4, value="used Depreciation")
        
        # Write active months
        phase_months = fitout_active_months[k] if k < len(fitout_active_months) else {}
        for c in range(6, 16):
            fy = start_year + (c - 6)
            col_char = openpyxl.utils.get_column_letter(c)
            
            # Write active months
            ws.cell(row=r_m, column=c, value=phase_months.get(fy, 0))
            # Write Annual Depreciation Rate formula
            ws.cell(row=r_c, column=c, value=f"=IF({col_char}{r_m}<>0,$B${r_c}/$B${r_d},0)")
            # Write used Depreciation formula
            ws.cell(row=r_d, column=c, value=f"={col_char}{r_c}/12*{col_char}{r_m}")

    # If N < 3, clear the labels and values of the unused template phases (from N to 3)
    if N < 3:
        for k in range(N, 3):
            r_m = 3 + 6 * k
            r_c = 6 + 6 * k
            r_d = 7 + 6 * k
            ws.cell(row=r_m, column=1, value="")
            ws.cell(row=r_m, column=4, value="")
            ws.cell(row=r_m, column=5, value="")
            ws.cell(row=r_m+1, column=4, value="")
            ws.cell(row=r_m+1, column=5, value="")
            ws.cell(row=r_c, column=1, value="")
            ws.cell(row=r_c, column=2, value=0.0)
            ws.cell(row=r_c, column=4, value="")
            ws.cell(row=r_d, column=1, value="")
            ws.cell(row=r_d, column=2, value="")
            ws.cell(row=r_d, column=4, value="")
            for c in range(6, 16):
                ws.cell(row=r_m, column=c, value=0)
                ws.cell(row=r_c, column=c, value=0)
                ws.cell(row=r_d, column=c, value=0)

    # Total Fitout Cost cell B3
    ws["B3"] = "=" + "+".join([f"B{6 + 6 * k}" for k in range(N)])
    
    # Calculate shifted row coordinates
    r_area_fit = 21 + offset
    r_area_mtr = 22 + offset
    r_inv_year = 23 + offset
    r_imp_int = 26 + offset
    r_used_dep = 27 + offset
    r_per_sqft = 29 + offset
    r_per_sqmtr = 30 + offset
    
    # Update Area, NBV, Interest, and Used Depreciation formulas for Fitout
    ws.cell(row=r_area_fit, column=4, value="Book value at Begin of FY")
    ws.cell(row=r_area_fit, column=6, value="=B3")
    for c in range(7, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        prev_col_char = openpyxl.utils.get_column_letter(c - 1)
        ws.cell(row=r_area_fit, column=c, value=f"={prev_col_char}{r_area_mtr}")
        
    ws.cell(row=r_area_mtr, column=4, value="Book value at End of FY")
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_area_mtr, column=c, value=f"={col_char}{r_area_fit}-{col_char}{r_used_dep}")
        
    ws.cell(row=r_inv_year, column=4, value="Average Net Book Value FY")
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_inv_year, column=c, value=f"=({col_char}{r_area_fit}+{col_char}{r_area_mtr})/2")
        
    ws.cell(row=r_imp_int, column=4, value="Impted Interest")
    ws.cell(row=r_imp_int, column=5, value=f"=SUM(F{r_imp_int}:O{r_imp_int})")
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_imp_int, column=c, value=f"={col_char}{r_inv_year}*$B$4/12*{col_char}3")
            
    ws.cell(row=r_used_dep, column=4, value="Used Depreciation")
    ws.cell(row=r_used_dep, column=5, value=f"=SUM(F{r_used_dep}:O{r_used_dep})")
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_used_dep, column=c, value="=" + "+".join([f"{col_char}{7 + 6 * k}" for k in range(N)]))
        
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_per_sqft, column=c, value=f"=SUM({col_char}{r_imp_int}:{col_char}{r_used_dep})")
        
    ws.cell(row=r_per_sqmtr, column=4, value="Monthly Value ")
    ws.cell(row=r_per_sqmtr, column=5, value=f"=IFERROR(SUM(F{r_imp_int}:O{r_imp_int},F{r_used_dep}:O{r_used_dep})/E3,0)")


    # ---------------- CAPEX SECTION ----------------
    r_capex_header = 36 + offset
    r_capex_cost = 37 + offset
    r_capex_val = 38 + offset
    
    # Write Capex schedule calendar years in Row r_capex_header
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_capex_header, column=c, value=f"={col_char}2")
        
    capex_sched = params.get("Capex Schedule", {})
    for c in range(6, 16):
        fy = start_year + (c - 6)
        ws.cell(row=r_capex_val, column=c, value=capex_sched.get(fy, 0.0))
        
    # Calculate and write active months for Capex tranches (Rows 40, 45, 50, 55, 60, 65, 70, 75, 80, 85 shifted)
    capex_lives = params.get("Capex Useful Lives", {})
    capex_active_months = params.get("Capex Active Months Breakdown", [])
    configured_years = list(capex_sched.keys()) + list(capex_lives.keys())
    if configured_years:
        num_capex_configured = min(10, max(configured_years) - start_year + 1)
    else:
        num_capex_configured = 0
        
    for i in range(1, 11):
        row_num = 40 + offset + 5 * (i - 1)
        yr = start_year + (i - 1)
        if i <= num_capex_configured:
            default_life = max(0, term_months - 12 * (i - 1))
            life = capex_lives.get(yr, default_life)
        else:
            life = 0
            
        # Write/Update formulas for the Capex tranche to handle shifted row references
        col_char_tranche = openpyxl.utils.get_column_letter(5 + i)
        ws.cell(row=row_num, column=2, value=f"=+{col_char_tranche}{38+offset}")
        ws.cell(row=row_num, column=5, value=f"=SUM(F{row_num}:O{row_num})")
        
        # Write useful life in years as a static number to avoid circular reference in SUMIF formulas
        ws.cell(row=row_num+1, column=2, value=life / 12.0)
        ws.cell(row=row_num+1, column=5, value=f"=E{row_num}-SUM(F{row_num}:N{row_num})")
        
        for c in range(6, 16):
            col_char = openpyxl.utils.get_column_letter(c)
            if i <= num_capex_configured:
                if c < 5 + i:
                    ws.cell(row=row_num, column=c, value=0)
                elif c == 5 + i:
                    ws.cell(row=row_num, column=c, value=f"=MIN($B${row_num+1}*12, 'Lease Rent'!$H$19)")
                else:
                    start_col = openpyxl.utils.get_column_letter(5 + i)
                    prev_col = openpyxl.utils.get_column_letter(c - 1)
                    ws.cell(row=row_num, column=c, value=f"=MIN(SUMIF('Lease Rent'!$A$31:$A$1048576, {col_char}${r_capex_header}, 'Lease Rent'!$D$31:$D$1048576), MAX(0, $B${row_num+1}*12 - SUM(${start_col}{row_num}:{prev_col}{row_num})))")
            else:
                ws.cell(row=row_num, column=c, value=0)
                
            # Write Annual Depreciation Rate formula
            ws.cell(row=row_num+2, column=c, value=f"=IF({col_char}{row_num}<>0,${col_char_tranche}${38+offset}/$B${row_num+1},0)")
            # Write used Depreciation formula
            ws.cell(row=row_num+3, column=c, value=f"={col_char}{row_num+2}/12*{col_char}{row_num}")
            
        # Show/Hide rows based on whether this tranche is configured
        for r in range(row_num, row_num + 5):
            ws.row_dimensions[r].hidden = (i > num_capex_configured)
            
    # Update Capex book value formulas (shifted by offset and offset_capex=9)
    r_cap_fit = 81 + offset + 9
    r_cap_mtr = 82 + offset + 9
    r_cap_year = 83 + offset + 9
    r_cap_imp = 86 + offset + 9
    r_cap_dep = 87 + offset + 9
    r_cap_per_sqft = 89 + offset + 9
    r_cap_per_sqmtr = 90 + offset + 9
    
    ws.cell(row=r_cap_fit, column=2, value=f"=B{21+offset}")
    ws.cell(row=r_cap_fit, column=6, value=f"=+F{38+offset}")
    for c in range(7, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        prev_col_char = openpyxl.utils.get_column_letter(c - 1)
        ws.cell(row=r_cap_fit, column=c, value=f"={prev_col_char}{r_cap_mtr}+{col_char}{38+offset}")
        
    for c in range(6, 16):
        col_char = openpyxl.utils.get_column_letter(c)
        ws.cell(row=r_cap_mtr, column=c, value=f"={col_char}{r_cap_fit}-{col_char}{r_cap_dep}")
        ws.cell(row=r_cap_year, column=c, value=f"=({col_char}{r_cap_fit}+{col_char}{r_cap_mtr})/2")
        ws.cell(row=r_cap_imp, column=c, value=f"={col_char}{r_cap_year}*$B${38+offset}/12*{col_char}{40+offset}")
        ws.cell(row=r_cap_dep, column=c, value="=" + "+".join([f"{col_char}{43 + offset + 5 * k}" for k in range(10)]))
        ws.cell(row=r_cap_per_sqft, column=c, value=f"=SUM({col_char}{r_cap_imp}:{col_char}{r_cap_imp+2})")
        
    ws.cell(row=r_cap_imp, column=5, value=f"=SUM(F{r_cap_imp}:O{r_cap_imp})")
    ws.cell(row=r_cap_dep, column=5, value=f"=SUM(F{r_cap_dep}:O{r_cap_dep})")
    ws.cell(row=r_cap_per_sqmtr, column=5, value=f"=IFERROR(SUM(F{r_cap_imp}:O{r_cap_dep})/E{40+offset},0)")

    # ---------------- PM SECTION ----------------
    r_pm_total = 95 + offset + 9
    r_pm_per_sqft = 96 + offset + 9
    r_pm_per_sqmtr = 97 + offset + 9
    
    # Write Maintenance (PM) schedule by FY
    pm_sched = params.get("PM Schedule", {})
    for c in range(6, 16):
        fy = start_year + (c - 6)
        ws.cell(row=r_pm_total, column=c, value=pm_sched.get(fy, 0.0))
        
    # Make PM total sum dynamic across all 10 years
    ws.cell(row=r_pm_total, column=2, value=f"=SUM(F{r_pm_total}:O{r_pm_total})")
    ws.cell(row=r_pm_per_sqft, column=2, value=f"=((B{r_pm_total}/B{21+offset}))/E3")
    ws.cell(row=r_pm_per_sqmtr, column=2, value=f"=B{r_pm_per_sqft}*10.764")

def simulate(params):
    """
    Simulates Capex & PM amortization schedules for UI preview.
    """
    start_date = params["Agreement Start Date"]
    start_year = start_date.year
    imputed_rate = params["Imputed Interest Rate"]
    term_months = params.get("Lease Term Months", 72)
    if pd.isna(term_months) or term_months is None:
        term_months = 72
        
    fitouts = list(params.get("Fitout Cost Breakdown", []))
    fitout_lifes = list(params.get("Fitout Useful Lives", []))
    fitout_active_months = params.get("Fitout Active Months Breakdown", [])
    
    N = len(fitouts)
    while len(fitouts) < 3:
        fitouts.append(0.0)
    while len(fitout_lifes) < 3:
        default_lifes = [term_months, 30, 48]
        fitout_lifes.append(default_lifes[len(fitout_lifes)])
    while len(fitout_active_months) < 3:
        fitout_active_months.append({})
        
    capex_sched = params.get("Capex Schedule", {})
    pm_sched = params.get("PM Schedule", {})
    
    years = list(range(start_year, start_year + 10))
    
    # Capex section calculations
    capex_active_months = params.get("Capex Active Months Breakdown", [])
    tranches = []
    capex_lives = params.get("Capex Useful Lives", {})
    for i in range(1, 11):
        yr = start_year + i - 1
        cost = capex_sched.get(yr, 0.0)
        default_life = max(0, term_months - 12 * (i - 1))
            
        life = capex_lives.get(yr, default_life)
        tranches.append({
            "cost": cost,
            "life": life,
            "start_year": yr
        })
        
    rows = []
    current_capex_book = 0.0
    
    for yr in years:
        # Calculate dynamic Fitout depreciation for all configured phases
        total_fitout_dep = 0.0
        active_f_months = 0
        for k in range(N):
            cost_k = fitouts[k]
            life_k = fitout_lifes[k]
            # Use custom months if present, otherwise default distribute
            phase_m = fitout_active_months[k] if k < len(fitout_active_months) else {}
            m_k = phase_m.get(yr, calculate_fy_months(start_date, max(term_months, 120), life_k).get(yr, 0))
            if k == 0:
                active_f_months = m_k  # use first phase active months for imputed interest active term reference
            dep_k = (cost_k / (life_k / 12) / 12) * m_k if m_k > 0 and life_k > 0 else 0.0
            total_fitout_dep += dep_k
            
        # Capex injected in current year
        capex_injected = capex_sched.get(yr, 0.0)
        book_begin = current_capex_book + capex_injected
        
        # Capex Depreciation
        capex_dep = 0.0
        for idx, t in enumerate(tranches):
            if yr >= t["start_year"] and t["life"] > 0:
                active_months = capex_active_months[idx].get(yr, 0) if idx < len(capex_active_months) and capex_active_months[idx] else get_capex_tranche_months(start_date, term_months, idx + 1, t["life"]).get(yr, 0)
                if active_months > 0:
                    capex_dep += (t["cost"] / t["life"]) * active_months
                    
        book_end = book_begin - capex_dep
        current_capex_book = book_end
        
        # Capex Imputed Interest
        active_term = calculate_fy_months(start_date, term_months, term_months).get(yr, 0)
        capex_imputed_interest = 0.0
        if active_term > 0:
            capex_imputed_interest = ((book_begin + book_end) / 2.0) * imputed_rate / 12.0 * active_term
            
        rows.append({
            "Fiscal Year": yr,
            "Fitout Months Active": active_f_months,
            "Fitout Amortization": round(total_fitout_dep, 2),
            "Capex Injected": round(capex_sched.get(yr, 0.0), 2),
            "Capex Amortization": round(capex_dep, 2),
            "Capex Imputed Interest": round(capex_imputed_interest, 2),
            "Maintenance Cost": round(pm_sched.get(yr, 0.0), 2)
        })
        
    return pd.DataFrame(rows)
